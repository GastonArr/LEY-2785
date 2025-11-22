import os
import shutil
import datetime as dt

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ==========================
# CONFIGURACIÓN BÁSICA
# ==========================
st.set_page_config(page_title="Carga LEY 2785", layout="wide")

TEMPLATE_FILE = "PLANILLA LEY 2785 NUEVA.xlsx"  # podés cambiar a .xlsm
EXCEL_SHEET_NAME = "LEY 2785"

# Mapear unidad -> archivo Excel de salida
UNIT_FILE_MAP = {
    "Comisaría 6°": "LEY-2785-COMISARIA6.xlsx",
    "Comisaría 9°": "LEY-2785-COMISARIA9.xlsx",
    "Comisaría 14°": "LEY-2785-COMISARIA14.xlsx",
    "Comisaría 15°": "LEY-2785-COMISARIA15.xlsx",
    "Comisaría 42°": "LEY-2785-COMISARIA42.xlsx",
    "CNAF 4": "LEY-2785-CNAF-4.xlsx",
}

UNIDADES_JURISDICCION = list(UNIT_FILE_MAP.keys())

# ==========================
# LISTAS DE OPCIONES
# ==========================
DOCUMENTO_OPTIONS = [
    "DNI/DU",
    "LC",
    "LE",
    "CI",
    "Pasaporte Extranjero",
    "Otro documento",
    "No informado",
]

SEXO1_OPTIONS = ["Mujer", "Varon", "No Informado"]
TRANS1_OPTIONS = ["Travesti", "Transexual", "Otra", "No informado"]

EDUCACION1_OPTIONS = [
    "Sin instrucción",
    "Primario",
    "EGB",
    "Polimodal",
    "Secundario",
    "Terciario",
    "Universitario",
    "Educación especial",
    "No sabe",
    "No informado",
]

COMPLITUD1_OPTIONS = ["SI", "NO", "NO SABE", "NO INFORMADO"]

OCUPADA1_OPTIONS = ["Ocupada/o", "No ocupada/o", "NO INFORMADO"]

ACTIVIDAD1_OPTIONS = [
    "estudiante",
    "jubilada/o-pensión",
    "Ama de casa",
    "otra",
    "No informado",
]

VINCULO_OPTIONS = [
    "Pareja/novio",
    "ex pareja",
    "padre",
    "madre",
    "hijo",
    "hija",
    "otros",
    "desconocido",
]

CONVIVENCIA_OPTIONS = ["si", "no", "no sabe", "no informado"]

TIPO_OPTIONS = ["SI", "NO"]

MODALIDAD_OPTIONS = [
    "Doméstica",
    "Institucional",
    "Laboral",
    "Contra la libertad reproductiva",
    "Obstétrica",
    "No informado",
]

TIEMPO_OPTIONS = [
    "Menos de un año",
    "de 1 a 5 años",
    "de 6 a 10 años",
    "más de 10 años",
    "No recureda",
    "No informado",
]

FRECUENCIA_OPTIONS = ["Sólo una vez", "Más de una vez", "No informado"]

SEXO2_OPTIONS = ["Mujer", "Varón", "No Sabe", "No informado"]
TRANS2_OPTIONS = ["Travesti", "Transexual", "Otra", "No informado"]

EDUCACION2_OPTIONS = list(EDUCACION1_OPTIONS)
COMPLITUD2_OPTIONS = ["Si", "No", "No Sabe", "No informado"]
ACTIVIDAD2_OPTIONS = ["Ocupada/o", "No ocupada/o", "No informado"]

OTRA2_OPTIONS = [
    "estudiante",
    "jubilada/o-pensión",
    "Ama de casa",
    "otra",
    "No informado",
]

# ==========================
# MAPEOS DE CAMPOS -> COLUMNAS
# ==========================
COLUMN_MAPPING = {
    "tipo_documento": "C",        # obligatorio
    "otro_doc": "D",              # NO obligatorio
    "identificacion": "E",        # obligatorio
    "institucion": "F",           # obligatorio
    "fecha_consulta": "G",        # obligatorio
    "sexo1": "H",                 # obligatorio
    "trans1": "I",                # obligatorio
    "edad": "J",                  # obligatorio
    "provincia": "K",             # obligatorio
    "partido_municipio": "L",     # NO obligatorio
    "localidad": "M",             # obligatorio
    "nivel_educativo1": "N",      # obligatorio
    "complitud1": "O",            # obligatorio
    "ocupada1": "P",              # obligatorio
    "actividad1": "Q",            # obligatorio
    "vinculo": "R",               # obligatorio
    "otro_vinculo": "S",          # NO obligatorio
    "convivencia": "T",           # obligatorio
    "viol_fisica": "U",           # obligatorio
    "viol_psico": "V",            # obligatorio
    "viol_econ": "W",             # obligatorio
    "viol_sexual": "X",           # obligatorio
    "modalidad": "Y",             # obligatorio
    "tiempo": "Z",                # obligatorio
    "frecuencia": "AA",           # obligatorio
    "sexo2": "AB",                # obligatorio
    "trans2": "AC",               # obligatorio
    "edad_agresor": "AD",         # obligatorio
    "nivel_educativo2": "AE",     # obligatorio
    "complitud2": "AF",           # obligatorio
    "actividad2": "AG",           # obligatorio
    "otra_actividad2": "AH",      # obligatorio
    "info_especifica": "AI",      # obligatorio
    "fecha_modificacion": "AJ",   # NO obligatorio
}

# Labels para mensajes de error
FIELD_LABELS = {
    "tipo_documento": "Tipo de documento (columna C)",
    "identificacion": "Identificación (columna E)",
    "institucion": "Unidad / Institución (columna F)",
    "fecha_consulta": "Fecha de consulta (columna G)",
    "sexo1": "Sexo (columna H)",
    "trans1": "Identidad trans (columna I)",
    "edad": "Edad (columna J)",
    "provincia": "Provincia (columna K)",
    "localidad": "Localidad (columna M)",
    "nivel_educativo1": "Nivel educativo (columna N)",
    "complitud1": "Complitud nivel educativo (columna O)",
    "ocupada1": "Situación ocupacional (columna P)",
    "actividad1": "Actividad (columna Q)",
    "vinculo": "Vínculo con el agresor (columna R)",
    "convivencia": "Convivencia (columna T)",
    "viol_fisica": "Violencia física (columna U)",
    "viol_psico": "Violencia psicológica (columna V)",
    "viol_econ": "Violencia económica (columna W)",
    "viol_sexual": "Violencia sexual (columna X)",
    "modalidad": "Modalidad de la violencia (columna Y)",
    "tiempo": "Tiempo del maltrato (columna Z)",
    "frecuencia": "Frecuencia de la violencia (columna AA)",
    "sexo2": "Sexo agresor (columna AB)",
    "trans2": "Identidad trans agresor (columna AC)",
    "edad_agresor": "Edad agresor (columna AD)",
    "nivel_educativo2": "Nivel educativo agresor (columna AE)",
    "complitud2": "Complitud educ. agresor (columna AF)",
    "actividad2": "Actividad laboral agresor (columna AG)",
    "otra_actividad2": "Otra actividad agresor (columna AH)",
    "info_especifica": "Información específica (columna AI)",
}

# Campos obligatorios por paso (para botón Siguiente)
STEP_REQUIRED = {
    1: ["institucion", "fecha_consulta"],
    2: [
        "tipo_documento",
        "identificacion",
        "sexo1",
        "trans1",
        "edad",
        "provincia",
        "localidad",
        "nivel_educativo1",
        "complitud1",
        "ocupada1",
        "actividad1",
    ],
    3: [
        "vinculo",
        "convivencia",
        "viol_fisica",
        "viol_psico",
        "viol_econ",
        "viol_sexual",
        "modalidad",
        "tiempo",
        "frecuencia",
    ],
    4: [
        "sexo2",
        "trans2",
        "edad_agresor",
        "nivel_educativo2",
        "complitud2",
        "actividad2",
        "otra_actividad2",
        "info_especifica",
    ],
}

# Lista de todos los campos obligatorios (se usa al guardar)
REQUIRED_FIELDS = sorted({key for keys in STEP_REQUIRED.values() for key in keys})

# ==========================
# FUNCIONES AUXILIARES
# ==========================
def initialize_default_state():
    """Precarga valores por defecto para que el validador no los considere vacíos."""

    default_values = {
        "institucion": UNIDADES_JURISDICCION[0],
        "fecha_consulta": dt.date.today(),
        "tipo_documento": DOCUMENTO_OPTIONS[0],
        "otro_doc": "",
        "sexo1": SEXO1_OPTIONS[0],
        "trans1": TRANS1_OPTIONS[0],
        "edad": 0,
        "provincia": "",
        "localidad": "",
        "partido_municipio": "",
        "nivel_educativo1": EDUCACION1_OPTIONS[0],
        "complitud1": COMPLITUD1_OPTIONS[0],
        "ocupada1": OCUPADA1_OPTIONS[0],
        "actividad1": ACTIVIDAD1_OPTIONS[0],
        "vinculo": VINCULO_OPTIONS[0],
        "convivencia": CONVIVENCIA_OPTIONS[0],
        "viol_fisica": TIPO_OPTIONS[0],
        "viol_psico": TIPO_OPTIONS[0],
        "viol_econ": TIPO_OPTIONS[0],
        "viol_sexual": TIPO_OPTIONS[0],
        "modalidad": MODALIDAD_OPTIONS[0],
        "tiempo": TIEMPO_OPTIONS[0],
        "frecuencia": FRECUENCIA_OPTIONS[0],
        "sexo2": SEXO2_OPTIONS[0],
        "trans2": TRANS2_OPTIONS[0],
        "edad_agresor": 0,
        "nivel_educativo2": EDUCACION2_OPTIONS[0],
        "complitud2": COMPLITUD2_OPTIONS[0],
        "actividad2": ACTIVIDAD2_OPTIONS[0],
        "otra_actividad2": OTRA2_OPTIONS[0],
        "info_especifica": "",
        "otro_vinculo": "",
        "fecha_modificacion": "",
    }

    for key, value in default_values.items():
        st.session_state.setdefault(key, value)


def ensure_unit_file_exists(unidad: str) -> str:
    if unidad not in UNIT_FILE_MAP:
        raise ValueError(f"Unidad no reconocida: {unidad}")

    target_file = UNIT_FILE_MAP[unidad]

    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(
            f"No se encontró el archivo plantilla '{TEMPLATE_FILE}'. "
            f"Dejalo en la misma carpeta que este script."
        )

    if not os.path.exists(target_file):
        shutil.copyfile(TEMPLATE_FILE, target_file)

    return target_file


def get_next_row_and_counter(ws):
    row = 3
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1

    if row == 3:
        next_number = 1
    else:
        last_val = ws.cell(row=row - 1, column=1).value
        try:
            last_val = int(last_val)
        except Exception:
            last_val = 0
        next_number = last_val + 1

    return row, next_number


def build_form_data_from_state():
    """Toma los valores del session_state y arma el dict que se escribirá en Excel."""
    sanitized_text = sanitize_required_text_fields()
    data = {}
    for key in COLUMN_MAPPING.keys():
        value = sanitized_text.get(key, st.session_state.get(key))

        if key == "fecha_consulta" and isinstance(value, dt.date):
            value = value.strftime("%d/%m/%Y")

        data[key] = value
    return data


def find_missing_in_state(keys):
    """Devuelve lista de claves que faltan (None o string vacío) leyendo directamente de session_state."""
    sanitized_text = sanitize_required_text_fields()

    missing = []
    for key in keys:
        val = sanitized_text.get(key, st.session_state.get(key, None))
        if val is None:
            missing.append(key)
        elif isinstance(val, str) and val.strip() == "":
            missing.append(key)
    return missing


def sanitize_required_text_fields():
    """Elimina espacios sobrantes y evita valores None en textos obligatorios."""

    # Incluimos partido_municipio para garantizar que se guarde la columna L.
    text_keys = ("identificacion", "provincia", "partido_municipio", "localidad")

    sanitized = {}
    for key in text_keys:
        val = st.session_state.get(key, "")
        if isinstance(val, str):
            val = val.strip()
        elif val is None:
            val = ""

        sanitized[key] = val
        # Actualizamos el session_state para que el guardado tome el valor saneado
        st.session_state[key] = val

    return sanitized


def save_to_excel(unidad, form_data):
    target_file = ensure_unit_file_exists(unidad)

    keep_vba = target_file.lower().endswith(".xlsm")
    wb = load_workbook(target_file, keep_vba=keep_vba)
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"La hoja '{EXCEL_SHEET_NAME}' no existe en el archivo {target_file}."
        )

    ws = wb[EXCEL_SHEET_NAME]

    row, counter = get_next_row_and_counter(ws)

    # Columna A: contador
    ws.cell(row=row, column=1).value = counter

    # Resto de columnas según mapping
    for key, col_letter in COLUMN_MAPPING.items():
        col_idx = column_index_from_string(col_letter)
        ws.cell(row=row, column=col_idx).value = form_data.get(key)

    wb.save(target_file)
    return counter, target_file


def reset_form():
    """Borra los datos del formulario (solo al finalizar guardado)."""
    for key in list(COLUMN_MAPPING.keys()):
        if key in st.session_state:
            del st.session_state[key]
    st.session_state.step = 1


# ==========================
# ESTADO DE LA APP
# ==========================
if "step" not in st.session_state:
    st.session_state.step = 1

initialize_default_state()

# ==========================
# UI - TÍTULO Y PROGRESO
# ==========================
st.title("Carga de registros - Ley 2785")

steps_total = 4
st.progress(st.session_state.step / steps_total)
st.caption(f"Paso {st.session_state.step} de {steps_total}")

# ==========================
# PASO 1
# ==========================
if st.session_state.step == 1:
    st.subheader("Paso 1: Unidad y datos básicos")

    st.selectbox(
        "Unidad / Institución (columna F)",
        UNIDADES_JURISDICCION,
        key="institucion",
        help="Solo se muestran las unidades de la jurisdicción.",
    )

    st.date_input(
        "Fecha de consulta (columna G)",
        key="fecha_consulta",
        help="Se guardará con el formato dd/mm/aaaa.",
        value=dt.date.today(),
    )

# ==========================
# PASO 2
# ==========================
elif st.session_state.step == 2:
    st.subheader("Paso 2: Datos de la persona consultante")

    st.selectbox(
        "Tipo de documento (columna C)",
        DOCUMENTO_OPTIONS,
        key="tipo_documento",
    )
    st.text_input(
        "Otro documento (columna D) [NO obligatorio]",
        key="otro_doc",
        value=st.session_state.get("otro_doc", ""),
    )
    st.text_input(
        "Identificación / N° doc (columna E)",
        key="identificacion",
        value=st.session_state.get("identificacion", ""),
    )

    cols1 = st.columns(2)
    with cols1[0]:
        st.selectbox("Sexo (columna H)", SEXO1_OPTIONS, key="sexo1")
    with cols1[1]:
        st.selectbox("Identidad trans (columna I)", TRANS1_OPTIONS, key="trans1")

    cols2 = st.columns(3)
    with cols2[0]:
        st.number_input(
            "Edad (columna J)",
            min_value=0,
            max_value=120,
            step=1,
            key="edad",
        )
    with cols2[1]:
        st.text_input(
            "Provincia (columna K)",
            key="provincia",
            value=st.session_state.get("provincia", ""),
        )
    with cols2[2]:
        st.text_input(
            "Partido / Municipio (columna L) [NO obligatorio]",
            key="partido_municipio",
            value=st.session_state.get("partido_municipio", ""),
        )

    cols3 = st.columns(3)
    with cols3[0]:
        st.text_input(
            "Localidad (columna M)",
            key="localidad",
            value=st.session_state.get("localidad", ""),
        )
    with cols3[1]:
        st.selectbox(
            "Nivel educativo (columna N)",
            EDUCACION1_OPTIONS,
            key="nivel_educativo1",
        )
    with cols3[2]:
        st.selectbox(
            "Complitud del nivel educativo (columna O)",
            COMPLITUD1_OPTIONS,
            key="complitud1",
        )

    cols4 = st.columns(2)
    with cols4[0]:
        st.selectbox(
            "Situación ocupacional (columna P)",
            OCUPADA1_OPTIONS,
            key="ocupada1",
        )
    with cols4[1]:
        st.selectbox(
            "Actividad (columna Q)",
            ACTIVIDAD1_OPTIONS,
            key="actividad1",
        )

# ==========================
# PASO 3
# ==========================
elif st.session_state.step == 3:
    st.subheader("Paso 3: Datos de la situación de violencia")

    st.selectbox(
        "Vínculo con el agresor (columna R)",
        VINCULO_OPTIONS,
        key="vinculo",
    )
    st.text_input(
        "Otro vínculo (columna S) [NO obligatorio]",
        key="otro_vinculo",
    )

    st.selectbox(
        "Convivencia con el agresor (columna T)",
        CONVIVENCIA_OPTIONS,
        key="convivencia",
    )

    cols1 = st.columns(4)
    with cols1[0]:
        st.selectbox(
            "Violencia física (columna U)",
            TIPO_OPTIONS,
            key="viol_fisica",
        )
    with cols1[1]:
        st.selectbox(
            "Violencia psicológica (columna V)",
            TIPO_OPTIONS,
            key="viol_psico",
        )
    with cols1[2]:
        st.selectbox(
            "Violencia económica (columna W)",
            TIPO_OPTIONS,
            key="viol_econ",
        )
    with cols1[3]:
        st.selectbox(
            "Violencia sexual (columna X)",
            TIPO_OPTIONS,
            key="viol_sexual",
        )

    cols2 = st.columns(3)
    with cols2[0]:
        st.selectbox(
            "Modalidad de la violencia (columna Y)",
            MODALIDAD_OPTIONS,
            key="modalidad",
        )
    with cols2[1]:
        st.selectbox(
            "Tiempo del maltrato (columna Z)",
            TIEMPO_OPTIONS,
            key="tiempo",
        )
    with cols2[2]:
        st.selectbox(
            "Frecuencia de la violencia (columna AA)",
            FRECUENCIA_OPTIONS,
            key="frecuencia",
        )

# ==========================
# PASO 4
# ==========================
elif st.session_state.step == 4:
    st.subheader("Paso 4: Datos del agresor y observaciones")

    cols1 = st.columns(3)
    with cols1[0]:
        st.selectbox(
            "Sexo agresor (columna AB)",
            SEXO2_OPTIONS,
            key="sexo2",
        )
    with cols1[1]:
        st.selectbox(
            "Identidad trans agresor (columna AC)",
            TRANS2_OPTIONS,
            key="trans2",
        )
    with cols1[2]:
        st.number_input(
            "Edad agresor (columna AD)",
            min_value=0,
            max_value=120,
            step=1,
            key="edad_agresor",
        )

    cols2 = st.columns(3)
    with cols2[0]:
        st.selectbox(
            "Nivel educativo agresor (columna AE)",
            EDUCACION2_OPTIONS,
            key="nivel_educativo2",
        )
    with cols2[1]:
        st.selectbox(
            "Complitud nivel educativo agresor (columna AF)",
            COMPLITUD2_OPTIONS,
            key="complitud2",
        )
    with cols2[2]:
        st.selectbox(
            "Actividad laboral agresor (columna AG)",
            ACTIVIDAD2_OPTIONS,
            key="actividad2",
        )

    st.selectbox(
        "Otra actividad agresor (columna AH)",
        OTRA2_OPTIONS,
        key="otra_actividad2",
    )

    st.text_area(
        "Información específica (columna AI)",
        key="info_especifica",
        height=150,
    )

    st.text_input(
        "Fecha de modificación (columna AJ) [NO obligatorio]",
        key="fecha_modificacion",
        help="Podés dejarlo vacío o usar el formato dd/mm/aaaa si corresponde.",
    )

    st.markdown("---")
    if st.button("💾 Guardar registro"):
        try:
            # Validamos únicamente los campos del paso 4
            required_step_keys = STEP_REQUIRED.get(4, [])
            missing_keys = find_missing_in_state(required_step_keys)

            if missing_keys:
                labels = [FIELD_LABELS[k] for k in missing_keys]
                st.error(
                    "Faltan completar los siguientes campos del paso 4:\n\n- "
                    + "\n- ".join(labels)
                )
            else:
                form_data = build_form_data_from_state()
                unidad = form_data.get("institucion")
                counter, filename = save_to_excel(unidad, form_data)
                st.success(
                    f"Registro guardado correctamente.\n\n"
                    f"Nº (columna A): **{counter}**\n"
                    f"Archivo: **{filename}**"
                )
                # limpiar formulario para que NO recuerde datos después de finalizar
                reset_form()
                st.rerun()
        except Exception as e:
            st.error(f"Error al guardar: {e}")

# ==========================
# BOTONES ANTERIOR / SIGUIENTE
# ==========================
st.markdown("---")
cols_nav = st.columns(3)

with cols_nav[0]:
    if st.session_state.step > 1:
        if st.button("⬅ Anterior"):
            st.session_state.step -= 1
            st.rerun()

with cols_nav[2]:
    if st.session_state.step < steps_total:
        if st.button("Siguiente ➡"):
            # Validamos campos del paso actual antes de avanzar
            required_step_keys = STEP_REQUIRED.get(st.session_state.step, [])
            missing_step = find_missing_in_state(required_step_keys)

            if missing_step:
                labels = [FIELD_LABELS[k] for k in missing_step]
                st.error(
                    "Para continuar, completá los siguientes campos:\n\n- "
                    + "\n- ".join(labels)
                )
            else:
                st.session_state.step += 1
                st.rerun()
